# inventario/views_reportes.py
import csv, json
from datetime import datetime, date, timedelta
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, F
from django.utils.timezone import make_naive

from .models import Venta, Sucursal, Producto
from .permissions import role_and_sucursales

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None  # Si no está instalado, mostramos instrucción en la vista


def _parse_date(value, default=None):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return default


def _permitted_sucursales(request):
    """Devuelve (rol, queryset_sucursales_permitidas, ids_permitidos_o_None)"""
    rol, sucursales = role_and_sucursales(request.user)
    if rol in ['Administrador', 'Subadministrador']:
        return rol, Sucursal.objects.all(), None  # None significa sin restricción
    ids = [s.id for s in sucursales]
    return rol, Sucursal.objects.filter(id__in=ids), ids


@login_required
def reportes_home(request):
    """Pantalla simple con filtros y enlace a CSV/Excel/Dashboard."""
    rol, qs_suc, _ids = _permitted_sucursales(request)
    hoy = date.today()
    ctx = {
        'sucursales': qs_suc,
        'hoy': hoy,
        'excel_disponible': bool(openpyxl),
    }
    return render(request, "reportes/reporte_ventas.html", ctx)


@login_required
def reporte_ventas_csv(request):
    """
    Exporta ventas a CSV: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&sucursal=<id|all>
    Respeta permisos por rol.
    """
    rol, _qs_suc, permitidas_ids = _permitted_sucursales(request)

    desde = _parse_date(request.GET.get("desde"), default=date.today())
    hasta = _parse_date(request.GET.get("hasta"), default=date.today())
    sucursal_param = request.GET.get("sucursal", "all")

    qs = Venta.objects.select_related("producto", "caja", "usuario", "caja__sucursal")
    qs = qs.filter(creado_en__date__gte=desde, creado_en__date__lte=hasta)

    if sucursal_param != "all":
        try:
            suc_id = int(sucursal_param)
        except ValueError:
            raise PermissionDenied("Parámetro de sucursal inválido.")
        if permitidas_ids is not None and suc_id not in permitidas_ids:
            raise PermissionDenied("No tienes permiso para esta sucursal.")
        qs = qs.filter(caja__sucursal_id=suc_id)
    else:
        if permitidas_ids is not None:
            qs = qs.filter(caja__sucursal_id__in=permitidas_ids)

    filename = f"ventas_{desde.isoformat()}_a_{hasta.isoformat()}.csv"
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(resp)
    writer.writerow([
        "Fecha/Hora", "Sucursal", "Caja ID", "Producto", "Cantidad",
        "Precio Unitario", "Total", "Usuario"
    ])

    total_general = 0
    for v in qs.order_by("creado_en"):
        writer.writerow([
            make_naive(v.creado_en).strftime("%Y-%m-%d %H:%M"),
            v.caja.sucursal.nombre,
            v.caja_id,
            v.producto.nombre,
            v.cantidad,
            f"{v.precio_unitario}",
            f"{v.total}",
            getattr(v.usuario, "username", "-"),
        ])
        total_general += float(v.total)

    writer.writerow([])
    writer.writerow(["", "", "", "", "", "TOTAL", f"{total_general:.2f}", ""])
    return resp


@login_required
def reporte_ventas_excel(request):
    """
    Exporta ventas a Excel (XLSX): ?desde=&hasta=&sucursal=
    Requiere openpyxl: pip install openpyxl
    """
    if openpyxl is None:
        # Mensaje claro si falta la librería
        return HttpResponse(
            "Para exportar a Excel instala openpyxl: pip install openpyxl",
            status=501, content_type="text/plain; charset=utf-8"
        )

    rol, _qs_suc, permitidas_ids = _permitted_sucursales(request)

    desde = _parse_date(request.GET.get("desde"), default=date.today())
    hasta = _parse_date(request.GET.get("hasta"), default=date.today())
    sucursal_param = request.GET.get("sucursal", "all")

    qs = Venta.objects.select_related("producto", "caja", "usuario", "caja__sucursal")
    qs = qs.filter(creado_en__date__gte=desde, creado_en__date__lte=hasta)

    if sucursal_param != "all":
        try:
            suc_id = int(sucursal_param)
        except ValueError:
            raise PermissionDenied("Parámetro de sucursal inválido.")
        if permitidas_ids is not None and suc_id not in permitidas_ids:
            raise PermissionDenied("No tienes permiso para esta sucursal.")
        qs = qs.filter(caja__sucursal_id=suc_id)
    else:
        if permitidas_ids is not None:
            qs = qs.filter(caja__sucursal_id__in=permitidas_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["Fecha/Hora", "Sucursal", "Caja ID", "Producto", "Cantidad", "Precio Unitario", "Total", "Usuario"]
    ws.append(headers)

    total_general = 0
    for v in qs.order_by("creado_en"):
        row = [
            make_naive(v.creado_en).strftime("%Y-%m-%d %H:%M"),
            v.caja.sucursal.nombre,
            v.caja_id,
            v.producto.nombre,
            v.cantidad,
            float(v.precio_unitario),
            float(v.total),
            getattr(v.usuario, "username", "-"),
        ]
        total_general += float(v.total)
        ws.append(row)

    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL", total_general, ""])

    # Ajuste básico de ancho
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    filename = f"ventas_{desde.isoformat()}_a_{hasta.isoformat()}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def reportes_dashboard(request):
    """
    Dashboard con dos gráficas:
      - Ventas por día (últimos 30 días)
      - Ventas por producto (top 10 en el rango)
    Filtros opcionales: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&sucursal=<id|all>
    """
    rol, qs_suc, permitidas_ids = _permitted_sucursales(request)

    # Rango por defecto: últimos 30 días
    hoy = date.today()
    default_desde = hoy - timedelta(days=29)
    desde = _parse_date(request.GET.get("desde"), default=default_desde)
    hasta = _parse_date(request.GET.get("hasta"), default=hoy)
    sucursal_param = request.GET.get("sucursal", "all")

    qs = Venta.objects.select_related("producto", "caja", "caja__sucursal")
    qs = qs.filter(creado_en__date__gte=desde, creado_en__date__lte=hasta)

    if sucursal_param != "all":
        try:
            suc_id = int(sucursal_param)
        except ValueError:
            raise PermissionDenied("Parámetro de sucursal inválido.")
        if permitidas_ids is not None and suc_id not in permitidas_ids:
            raise PermissionDenied("No tienes permiso para esta sucursal.")
        qs = qs.filter(caja__sucursal_id=suc_id)
    else:
        if permitidas_ids is not None:
            qs = qs.filter(caja__sucursal_id__in=permitidas_ids)

    # Serie diaria
    dias = [default_desde + timedelta(days=i) for i in range((hasta - default_desde).days + 1)]
    mapa_dias = {d: 0.0 for d in dias}

    agreg_dia = (
        qs.values("creado_en__date")
          .annotate(total=Sum("total"))
    )
    for item in agreg_dia:
        d = item["creado_en__date"]
        if d in mapa_dias:
            mapa_dias[d] = float(item["total"] or 0)

    labels_dias = [d.strftime("%Y-%m-%d") for d in dias]
    data_dias = [mapa_dias[d] for d in dias]

    # Top productos (por total vendido)
    agreg_prod = (
        qs.values("producto__nombre")
          .annotate(total=Sum("total"))
          .order_by("-total")[:10]
    )
    labels_prod = [x["producto__nombre"] for x in agreg_prod]
    data_prod = [float(x["total"] or 0) for x in agreg_prod]

    ctx = {
        "sucursales": qs_suc,
        "desde": desde,
        "hasta": hasta,
        "labels_dias": json.dumps(labels_dias),
        "data_dias": json.dumps(data_dias),
        "labels_prod": json.dumps(labels_prod),
        "data_prod": json.dumps(data_prod),
        "sucursal_param": sucursal_param,
        "hoy": hoy,
    }
    return render(request, "reportes/dashboard.html", ctx)
